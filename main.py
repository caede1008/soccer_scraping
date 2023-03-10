import os
import sys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
from selenium.webdriver.common.by import By
import openpyxl

#　エクセル設定
path = r'C:/Users/uu/soccer_scraping_excel/players.xlsx'
wb = openpyxl.load_workbook(path)

# 変数宣言
sheetcnt = 1
url = ''

try:
    sheetcnt = len(wb.worksheets)
    if sheetcnt > 100:
        raise Exception
    for sheet in range (1, sheetcnt + 1):
        ws = wb[str(sheet)]
        # エクセルからURLを取得
        url = ws['B1'].value
        if url == "":
            raise Exception

        # ドライバー起動
        driver = webdriver.Chrome(r'C:/chromedriver')
        driver.get(url)
        time.sleep(3)

        # スクレイピング A1
        playername = driver.find_element(By.CLASS_NAME, "sc-bqWxrE.fxLCLd").text
        club = driver.find_element(By.CLASS_NAME, "sc-bqWxrE.jLbhyz").text
        nationality = driver.find_element(By.CLASS_NAME, "sc-hLBbgP.sc-eDvSVe.gjJmZQ.kPTbfh").text
        wkdata = driver.find_elements(By.CLASS_NAME, "sc-bqWxrE.hBBdLz")
        year = wkdata[1].text
        height = wkdata[2].text
        foot = wkdata[3].text
        position = wkdata[4].text
        shirt_number = wkdata[5].text

        # スクレイピング A2

        # 前回入力値クリアー
        for rows in ws['A12':'F40']:
            for cell in rows:
                cell.value = None

        leagues_games = driver.find_elements(By.XPATH,
                                             "//*[@id='__next']/div/main/div[1]/div/div[1]/div[3]/div/div[2]/div[1]/div/div[2]/*")

        gamedates = []
        wkgamedates = driver.find_elements(By.CLASS_NAME, "sc-bqWxrE.gffDkV")
        for wkgamedate in wkgamedates:
            if len(wkgamedate.text) != 0:
                gamedates.append(wkgamedate.text)

        gamepositions = []
        wkgamepositions = driver.find_elements(By.CLASS_NAME,
                                               "sc-hLBbgP.sc-eDvSVe.fuUKnP.hyKYsT.sc-9199a964-2.kgwLqG.score-box")
        for wkgameposition in wkgamepositions:
            if len(wkgameposition.text) != 0:
                gamepositions.append(wkgameposition.text)

        wkmatches = driver.find_elements(By.CLASS_NAME, 'sc-hLBbgP.eIlfTT')
        matches = []
        for wkmatche in wkmatches:
            data_title = wkmatche.get_attribute("title")
            matches.append(data_title)

        wkselfscores = driver.find_elements(By.CLASS_NAME,
                                            "sc-hLBbgP.sc-eDvSVe.fuUKnP.bMwHQt.sc-9199a964-2.kgwLqG.score-box")
        selfscores = []
        for wkselfscore in wkselfscores:
            if len(wkselfscore.text) != 0:
                selfscores.append(wkselfscore.text)
        selfscores += ["", ""]

        wkrates = driver.find_elements(By.CLASS_NAME, "sc-bqWxrE.gGeeTx")
        rates = []
        for wkrate in wkrates:
            if len(wkrate.text) != 0:
                rates.append(wkrate.text)
            else:
                rates.append("")

        # A1エクセル書き込み
        ws.cell(3, 2).value = playername
        ws.cell(4, 2).value = club
        ws.cell(5, 2).value = nationality
        ws.cell(6, 2).value = year
        ws.cell(7, 2).value = height
        ws.cell(8, 2).value = foot
        ws.cell(9, 2).value = position
        ws.cell(10, 2).value = shirt_number

        # A2エクセル書き込み
        idxcnt = 0
        scoreidx = 0
        rownumber = 12
        for league_game in leagues_games:
            type = league_game.get_attribute("class")
            # リーグの場合
            if type == 'sc-hLBbgP sc-eDvSVe jhoWjm fRddxb':
                rownumber += 1
                ws.cell(rownumber, 1).value = league_game.text
            else:
                rownumber += 1
                ws.cell(rownumber, 2).value = gamedates[idxcnt]
                ws.cell(rownumber, 3).value = gamepositions[idxcnt]
                ws.cell(rownumber, 4).value = matches[idxcnt]
                ws.cell(rownumber, 5).value = selfscores[scoreidx] + '-' + selfscores[scoreidx + 1]
                ws.cell(rownumber, 6).value = rates[idxcnt]
                idxcnt += 1
                scoreidx += 2

        driver.close()
        wb.save(path)
        sheetcnt += 1

    wb.close()

except:
    sheetnum = str(sheetcnt)
    print("シート" + sheetnum + "にてエラーが発生しました。")
    wb.close()


